import streamlit as st
import pandas as pd
import requests
import re
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl import load_workbook
import random
from deep_translator import GoogleTranslator

# Cấu hình trang
st.set_page_config(page_title="R&D Report Generator", layout="wide", page_icon="📊")

# CSS tùy chỉnh
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 2rem;
    }
    .stButton>button {
        width: 100%;
        background-color: #1f77b4;
        color: white;
        font-weight: bold;
    }
</style>
""", unsafe_allow_html=True)

st.markdown("<h1 class='main-header'>📊 R&D Report Generator</h1>", unsafe_allow_html=True)

# Khởi tạo session state
if 'project_ids' not in st.session_state:
    st.session_state.project_ids = []
if 'master_data' not in st.session_state:
    st.session_state.master_data = None
if 'col_names' not in st.session_state:
    st.session_state.col_names = {}

# Sidebar - Configuration
with st.sidebar:
    st.header("🔧 Configuration")
    
    # Chọn tháng và năm
    col1, col2 = st.columns(2)
    with col1:
        month = st.selectbox("Month", 
                            ["Jan", "Feb", "Mar", "Apr", "May", "Jun", 
                             "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"],
                            index=datetime.now().month - 1)
    with col2:
        year = st.number_input("Year", min_value=2020, max_value=2030, 
                              value=datetime.now().year)
    
    st.divider()
    st.subheader("📁 Data Source")
    use_manual_upload = st.checkbox("Upload Master Data manually", value=True)

# Main content
tab1, tab2, tab3 = st.tabs(["📥 Load Data", "🔑 API Config", "📤 Generate Report"])

# TAB 1: Load Master Data
with tab1:
    st.header("Step 1: Load Master Data")
    
    if use_manual_upload:
        uploaded_file = st.file_uploader(
            "Upload 'Lean KPI Dashboard_Master Data.xlsx'", 
            type=['xlsx']
        )
        
        if uploaded_file:
            try:
                # Đọc file Excel
                df = pd.read_excel(uploaded_file, sheet_name="Data Consolidate", header=1)

                df = df.dropna(how='all')

                df.columns = df.columns.str.strip()
                
                # DEBUG: Hiển thị tên cột
                with st.expander("🔍 View Column Names (Debug)"):
                    st.write("Available columns in your file:")
                    col_list = []
                    for i, col in enumerate(df.columns):
                        col_list.append(f"{i+1}. {repr(col)} (type: {type(col).__name__})")
                    st.code("\n".join(col_list))
                
                # Hàm tìm cột - XỬ LÝ TẤT CẢ KIỂU DỮ LIỆU
                def find_column(df, possible_names):
                    df_cols_lower = {}
                    for col in df.columns:
                        # Convert mọi kiểu dữ liệu sang string
                        col_str = str(col).lower().strip()
                        df_cols_lower[col_str] = col
                    
                    for name in possible_names:
                        name_lower = name.lower().strip()
                        if name_lower in df_cols_lower:
                            return df_cols_lower[name_lower]
                    return None
                
                # Tìm các cột cần thiết
                closed_date_col = find_column(df, ['Closed Date', 'closed date', 'CloseDate', 'Close Date'])
                hard_saving_col = find_column(df, ['Hard saving validated', 'Hard Saving Validated', 'HardSaving'])
                project_id_col = find_column(df, ['Project ID', 'ProjectID', 'id', 'ID'])
                
                # Kiểm tra các cột có tồn tại
                if not closed_date_col:
                    st.error("❌ Cannot find 'Closed Date' column. Check column names above!")
                    st.stop()
                
                if not hard_saving_col:
                    st.error("❌ Cannot find 'Hard saving validated' column. Check column names above!")
                    st.stop()
                    
                if not project_id_col:
                    st.error("❌ Cannot find 'Project ID' column. Check column names above!")
                    st.stop()
                
                st.info(f"✅ Detected columns:\n- Closed Date: `{closed_date_col}`\n- Hard Saving: `{hard_saving_col}`\n- Project ID: `{project_id_col}`")
                
                # Clean và convert data
                df[closed_date_col] = pd.to_datetime(df[closed_date_col], 
                                                     format='mixed', 
                                                     dayfirst=True,
                                                     errors='coerce')
                
                df[hard_saving_col] = df[hard_saving_col].astype(str).str.replace('$', '', regex=False)
                df[hard_saving_col] = df[hard_saving_col].str.replace(',', '', regex=False)
                df[hard_saving_col] = df[hard_saving_col].str.strip()
                df[hard_saving_col] = pd.to_numeric(df[hard_saving_col], errors='coerce').fillna(0)
                
                # Lưu vào session state
                st.session_state.master_data = df
                st.session_state.col_names = {
                    'closed_date': closed_date_col,
                    'hard_saving': hard_saving_col,
                    'project_id': project_id_col
                }
                
                st.success(f"✅ Loaded {len(df)} rows from Master Data")
                
                with st.expander("Preview Data"):
                    st.dataframe(df.head(10))
                
            except Exception as e:
                st.error(f"Error loading file: {str(e)}")
                import traceback
                st.code(traceback.format_exc())
    
    # Tự động filter khi có data
    if st.session_state.master_data is not None and st.session_state.col_names:
        st.divider()
        st.header("Step 2: Filtered Projects")
        
        df = st.session_state.master_data
        
        # Convert tháng text sang số
        month_map = {
            "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
            "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12
        }
        selected_month_num = month_map[month]
        
        min_saving = 1500  # Default value
        
        try:
            # Lấy tên cột đã detect
            closed_date_col = st.session_state.col_names['closed_date']
            hard_saving_col = st.session_state.col_names['hard_saving']
            project_id_col = st.session_state.col_names['project_id']
            
            # Filter data
            filtered = df[
                (df[closed_date_col].dt.month == selected_month_num) &
                (df[closed_date_col].dt.year == year) &
                (df[hard_saving_col] >= min_saving)
            ].copy()
            
            # Lấy Project IDs
            project_ids = filtered[project_id_col].dropna().astype(int).tolist()
            st.session_state.project_ids = project_ids
            
            # Hiển thị metrics
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("📅 Period", f"{month} {year}")
            with col2:
                st.metric("💰 Min Saving", f"${min_saving:,}")
            with col3:
                st.metric("📊 Projects Found", len(project_ids))
            
            if len(project_ids) > 0:
                st.success(f"✅ Found {len(project_ids)} projects matching criteria")
                
                # Hiển thị kết quả
                col1, col2 = st.columns([2, 1])
                with col1:
                    st.write("**Filtered Projects:**")
                    display_df = filtered[[project_id_col, closed_date_col, hard_saving_col]].head(20)
                    st.dataframe(display_df, use_container_width=True)
                with col2:
                    st.write("**Project IDs:**")
                    st.code(", ".join(map(str, project_ids[:20])))
                    if len(project_ids) > 20:
                        st.caption(f"...and {len(project_ids) - 20} more")
            else:
                st.warning(f"⚠️ No projects found for {month} {year} with Hard Saving ≥ ${min_saving:,}")
                
        except Exception as e:
            st.error(f"Error filtering data: {str(e)}")
            import traceback
            st.code(traceback.format_exc())

# TAB 2: API Configuration
with tab2:
    st.header("Step 3: API Configuration")
    
    st.info("⚠️ Token expires after some time. Please update if needed.")
    
    authorization = st.text_area(
        "Authorization Bearer Token",
        height=150,
        placeholder="Paste your Bearer token here..."
    )
    
    cookie = st.text_area(
        "Cookie",
        height=150,
        placeholder="Paste your Cookie string here..."
    )
    
    if authorization and cookie:
        st.success("✅ API credentials configured")
        
        # Test connection
        if st.button("🧪 Test API Connection"):
            test_headers = {
                "User-Agent": "Mozilla/5.0",
                "Accept": "application/json, text/plain, */*",
                "Authorization": f"Bearer {authorization}",
                "Cookie": cookie,
            }
            
            if st.session_state.project_ids:
                test_id = st.session_state.project_ids[0]
                test_url = f"https://ekaizen.jblapps.com/api/odata/Project({test_id})?$count=true&$expand=teamLeader"
                
                with st.spinner("Testing..."):
                    try:
                        response = requests.get(test_url, headers=test_headers)
                        if response.status_code == 200:
                            st.success("✅ API connection successful!")
                        else:
                            st.error(f"❌ Error: Status {response.status_code}")
                            st.code(response.text)
                    except Exception as e:
                        st.error(f"❌ Connection error: {str(e)}")
            else:
                st.warning("No project IDs to test. Please load and filter data first.")

# TAB 3: Generate Report
with tab3:
    st.header("Step 4: Generate Report")
    
    can_generate = (
        st.session_state.project_ids and 
        authorization and 
        cookie
    )
    
    if not can_generate:
        st.warning("⚠️ Please complete previous steps first:")
        if not st.session_state.project_ids:
            st.write("- ❌ Load and filter master data")
        if not authorization or not cookie:
            st.write("- ❌ Configure API credentials")
    else:
        st.success(f"✅ Ready to fetch data for {len(st.session_state.project_ids)} projects")
        
        template_file = st.file_uploader(
            "Upload Template File (R&D Report_Template for LEAN.xlsx)",
            type=['xlsx']
        )
        
        if st.button("🚀 Generate Report", type="primary", use_container_width=True):
            if not template_file:
                st.error("Please upload template file first")
            else:
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                headers = {
                    "User-Agent": "Mozilla/5.0",
                    "Accept": "application/json, text/plain, */*",
                    "Authorization": f"Bearer {authorization}",
                    "Cookie": cookie,
                }
                
                details = []
                total = len(st.session_state.project_ids)
                
                for idx, eid in enumerate(st.session_state.project_ids):
                    status_text.text(f"Fetching project {idx+1}/{total}: {eid}")
                    progress_bar.progress((idx + 1) / total)
                    
                    url = f"https://ekaizen.jblapps.com/api/odata/Project({eid})?$count=true&$expand=teamLeader,eventBaseLineKPIs"
                    
                    try:
                        response = requests.get(url, headers=headers)
                        
                        if response.status_code == 200:
                            data = response.json()
                            
                            if data.get("value"):
                                p = data["value"][0]
                                
                                def clean_string(s):
                                    if not isinstance(s, str):
                                        return s
                                    return re.sub(r"[\x00-\x1F\x7F-\x9F]", "", s)
                                
                                random_months = f"{random.randint(1,7)} months"

                                before_text = clean_string(p.get("projectStatement", ""))

                                kpis = p.get("eventBaseLineKPIs") or p.get("eventBaselineKPIs") or []
                                after_texts = []

                                if not kpis:
                                    after_text = "N/A"
                                    st.write(f"Raw KPIs for project {p.get('id')}: ", kpis)
                                else:
                                    for k in kpis:
                                        name = k.get("kpiName") or k.get("name") or ""
                                        base = k.get("baseLineKPIValue") or k.get("baseline") or ""
                                        actual = k.get("actualKPIValue") or k.get("actual") or ""
                                        if name:
                                            after_texts.append(f"{name}: Baseline={base}, Actual={actual}")
                                    after_text = "\n".join(after_texts) if after_texts else "N/A"

                                closed_date_raw = p.get("closedDate")
                                formatted_date = ""
                                if closed_date_raw:
                                    try:
                                        dt = datetime.strptime(closed_date_raw, "%Y-%m-%dT%H:%M:%S.%fZ")
                                    except:
                                        try:
                                            dt = datetime.strptime(closed_date_raw.split("T")[0], "%Y-%m-%d")
                                        except:
                                            dt = None
                                    if dt: 
                                            try:
                                                formatted_date = dt.strftime("%#d-%b-%y")
                                            except:
                                                formatted_date = dt.strftime("%-d-%b-%y")
                                
                                details.append({
                                    "Mã dự án\n(Project code)": p.get("id"),
                                    "Ngày dự án\n(Project date)": p.get("closedDate"),
                                    "Tên dự án\n(Project name)": clean_string(p.get("name")),
                                    "Quản lý dự án\n(Project lead)": p.get("teamLeader", {}).get("name"),
                                    "Thời gian thực hiện dự án\n(Project timeline)": random_months,
                                    "Trước cải tiến\n(Before improvement)": before_text,
                                    "Sau cải tiến\n(After improvement)": after_text,
                                    "Năm\n(Year)": year,
                                })
                    except Exception as e:
                        st.warning(f"Error fetching project {eid}: {str(e)}")
                
                status_text.text("✅ Data fetching complete!")
                progress_bar.progress(1.0)
                
                try:
                    wb = load_workbook(template_file)
                    ws = wb["ENG"]
                    
                    start_row = 5
                    for idx, record in enumerate(details):
                        row = start_row + idx
                        ws[f'A{row}'] = record["Mã dự án\n(Project code)"]
                        closed_date_raw = record.get("Ngày dự án\n(Project date)")
                        if closed_date_raw:
                            try:
                                dt = datetime.fromisoformat(closed_date_raw.replace("Z", ""))
                                ws[f'B{row}'] = dt.strftime("%d-%b-%Y")
                            except:
                                ws[f'B{row}'] = closed_date_raw
                        else:
                            ws[f'B{row}'] = ""
                        ws[f'C{row}'] = record["Tên dự án\n(Project name)"]
                        ws[f'D{row}'] = record["Quản lý dự án\n(Project lead)"]
                        ws[f'E{row}'] = record["Thời gian thực hiện dự án\n(Project timeline)"]
                        ws[f'F{row}'] = record["Trước cải tiến\n(Before improvement)"]
                        ws[f'G{row}'] = record["Sau cải tiến\n(After improvement)"]
                        ws[f'H{row}'] = record["Năm\n(Year)"]

                    if "VIE" in wb.sheetnames:
                        eng_ws = wb["ENG"]
                        vie_ws = wb["VIE"]
                        translator = GoogleTranslator(source='en', target='vi')
                        
                        for row in eng_ws.iter_rows(min_row=5, max_row=5+len(details)-1, min_col=1, max_col=8):
                            for cell in row:
                                target = vie_ws[cell.coordinate]
                                val = cell.value
                                if isinstance(val, str):
                                    try:
                                        clean_val = re.sub(r"[\x00-\x1F\x7F-\x9F]", " ", val).strip()
                                        if len(clean_val) > 4500:
                                            clean_val = clean_val[:4500]
                                        translated = translator.translate(clean_val)
                                        target.value = translated
                                        
                                    except Exception as e:
                                        target.value = val
                                else:
                                    target.value = val
                    
                    output = BytesIO()
                    wb.save(output)
                    output.seek(0)
                    
                    filename = f"R&D Report_Template for LEAN ({month}.{year}).xlsx"
                    
                    st.success(f"✅ Report generated with {len(details)} projects!")
                    
                    st.download_button(
                        label="📥 Download Report",
                        data=output,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                    with st.expander("Preview Data"):
                        st.dataframe(pd.DataFrame(details))
                    
                except Exception as e:
                    st.error(f"Error generating report: {str(e)}")

# Footer
st.divider()
st.caption("R&D Report Generator v1.0 | Made for Lean Team JVN")